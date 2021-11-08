import numpy as np
import random

# 問題の設定
PROCESS_MIN = 1 # 加工時間の最小
PROCESS_MAX = 10 # 加工時間の最大
JOB = 10 # ジョブの個数
WORKER = 3 # 作業員の人数
PATTERN = 3 # 部品タイプの数
MACHINE_SELECTION = {1:[1,2],2:[3,4],3:[5,6]} # タイプから担当機械を選択する辞書を作成
MACHINE = max(max(MACHINE_SELECTION.values()))
import openpyxl
from openpyxl.descriptors.base import Length
#ファイルのパスを指定
file_path = r"C:\Users\nabem\Documents\sample.xlsx"

#ファイルを開く
excelBook = openpyxl.load_workbook(file_path)

#特定のシートを読み込む
excelSheet = excelBook["Sheet1"]

#作業者と機械のペアのリストを作成
worker_machine = []
#作業者３人と機械６台のペア
# #作業者３人
for w in range(3,excelSheet.max_row+1):
    #機械６台
    for i in range(3,excelSheet.max_column+1):
        pair_worker_machine = (excelSheet.cell(row=w,column=2).value,excelSheet.cell(row=2,column=i).value)
        worker_machine.append(pair_worker_machine)

#技能度のリストを作成
skill = []
#作業者３人の機械6台における技能度
# #作業者３人
for v in range(3,excelSheet.max_row+1):
    #機械6台
    for z in range(3,excelSheet.max_column+1):
        SKILL_LEVEL_worker_machine = excelSheet.cell(row=3,column=z).value
        skill.append(SKILL_LEVEL_worker_machine)

#2つのリストから辞書型を生成
SKILL_LEVEL = dict(zip(worker_machine,skill))


class SimpleGA:
    def __init__(self):
        self.N = 100 # 個体数
        self.ITERATION = 100 # 世代数
        self._initialize()

    def _initialize(self):
        self.pool = [Chromosome() for i in range(self.N)]

    def _mutateAll(self): # 突然変異
        pass

    def _Selction(): # 選択
        pass

    def evolve(self):
        for i in range(self.ITERATION):
            self._Selction()
            self._mutateAll()
            pass

class Chromosome:
    def __init__(self):
        self.array_process_time = np.random.randint(PROCESS_MIN,PROCESS_MAX,(JOB,1)) # ジョブごとの作業時間の生成
        self.array_worker = np.random.randint(1,WORKER+1,(JOB,1)) # ジョブごとの担当作業者の決定
        self.machine_pattern = np.random.randint(1,PATTERN+1,(JOB,1)) # 部品タイプの決定
        # 部品タイプから担当できる機械を選択
        self.selected_machine = []
        for i in range(JOB):
            self.machine_dict = self.machine_pattern[i]
            self.machine_list = MACHINE_SELECTION[self.machine_dict[0]]
            self.sample = random.randint(min(self.machine_list),max(self.machine_list))
            self.selected_machine.append(self.sample)
        self.array_selected_machine = np.array(self.selected_machine).reshape(JOB,1)
        
        self.order_sample = [0] * MACHINE
        self.order = []
        
        for i in self.array_selected_machine:
            self.order_sample[int(i-1)] += 1
            self.order.append(self.order_sample[int(i-1)])
            
        self.array_order = np.array(self.order).reshape(JOB,1)
        self.list_net_work_time = []
        for i in range(JOB):
            self.ability = SKILL_LEVEL[(int(self.array_worker[i])),int(self.array_selected_machine[i])]
            self.net_process_time = self.array_process_time[i] * self.ability
            self.list_net_work_time.append(self.net_process_time)
        self.array_net_work_time = np.array(self.list_net_work_time).reshape(JOB,1)
        
        # 一つの行列に結合
        self.gene = np.concatenate([self.array_process_time,self.array_worker,self.array_selected_machine,
                                    self.array_order,self.array_net_work_time],axis = 1)
        
        # self.MUTATION_RATE = 0.05 # 突然変異率
    
    def mutate(self):
        pass

    def getVal(self): # 遺伝子型から表現型への発現
        # net_work_timeの関数を作成
        pass
    def getFitness(self): # 適応度評価
        pass